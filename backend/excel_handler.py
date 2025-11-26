from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from pathlib import Path
import os

EXCEL_FILE_PATH = "emprestimos_backup.xlsx"

def export_loans_to_excel(emprestimos, historicos=None):
    """
    Exporta empréstimos e histórico para arquivo Excel
    
    Args:
        emprestimos: Lista de objetos Emprestimo
        historicos: Lista de objetos HistoricoValorAdiantado (opcional)
    
    Returns:
        str: Caminho do arquivo Excel gerado
    """
    wb = Workbook()
    
    # Aba de Empréstimos
    ws_loans = wb.active
    ws_loans.title = "Empréstimos"
    
    # Cabeçalhos
    headers = [
        "ID", "Descrição", "Instituição Credora", "Valor Parcela", 
        "Valor Parcela Adiantada", "Qtd Total Parcelas", "Qtd Parcelas Devidas",
        "Taxa SELIC (%)", "Taxa CDI (%)", "Data Cadastro", "Dia Vencimento"
    ]
    
    # Estilizar cabeçalhos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_loans.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados dos empréstimos
    for row_num, emp in enumerate(emprestimos, 2):
        ws_loans.cell(row=row_num, column=1, value=emp.id)
        ws_loans.cell(row=row_num, column=2, value=emp.descricao)
        ws_loans.cell(row=row_num, column=3, value=emp.instituicao_credora)
        ws_loans.cell(row=row_num, column=4, value=emp.valor_parcela)
        ws_loans.cell(row=row_num, column=5, value=emp.valor_parcela_adiantada)
        ws_loans.cell(row=row_num, column=6, value=emp.qtd_total_parcelas)
        ws_loans.cell(row=row_num, column=7, value=emp.qtd_parcelas_devidas)
        ws_loans.cell(row=row_num, column=8, value=emp.taxa_selic_registro)
        ws_loans.cell(row=row_num, column=9, value=emp.taxa_cdi_registro)
        ws_loans.cell(row=row_num, column=10, value=emp.data_cadastro)
        ws_loans.cell(row=row_num, column=11, value=emp.dia_vencimento)
    
    # Ajustar largura das colunas
    for column in ws_loans.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_loans.column_dimensions[column_letter].width = adjusted_width
    
    # Aba de Histórico (se houver)
    if historicos:
        ws_history = wb.create_sheet("Histórico")
        
        history_headers = [
            "ID", "Empréstimo ID", "Data Registro", 
            "Valor Parcela Adiantada", "Taxa SELIC (%)", "Taxa CDI (%)"
        ]
        
        for col_num, header in enumerate(history_headers, 1):
            cell = ws_history.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row_num, hist in enumerate(historicos, 2):
            ws_history.cell(row=row_num, column=1, value=hist.id)
            ws_history.cell(row=row_num, column=2, value=hist.emprestimo_id)
            ws_history.cell(row=row_num, column=3, value=hist.data_registro)
            ws_history.cell(row=row_num, column=4, value=hist.valor_parcela_adiantada)
            ws_history.cell(row=row_num, column=5, value=hist.taxa_selic)
            ws_history.cell(row=row_num, column=6, value=hist.taxa_cdi)
        
        # Ajustar largura das colunas
        for column in ws_history.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_history.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar arquivo
    wb.save(EXCEL_FILE_PATH)
    return EXCEL_FILE_PATH


def auto_sync_to_excel(db):
    """
    Sincroniza automaticamente os dados do banco para o Excel
    
    Args:
        db: Sessão do banco de dados
    """
    from database import Emprestimo, HistoricoValorAdiantado
    
    try:
        emprestimos = db.query(Emprestimo).all()
        historicos = db.query(HistoricoValorAdiantado).all()
        
        export_loans_to_excel(emprestimos, historicos)
        print(f"✅ Sincronização automática realizada: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    except Exception as e:
        print(f"❌ Erro na sincronização automática: {str(e)}")


def import_loans_from_excel(file_path, db):
    """
    Importa empréstimos do arquivo Excel para o banco de dados
    
    Args:
        file_path: Caminho do arquivo Excel
        db: Sessão do banco de dados
    
    Returns:
        dict: Resultado da importação com contadores
    """
    from database import Emprestimo, HistoricoValorAdiantado
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
    
    wb = load_workbook(file_path)
    
    loans_imported = 0
    history_imported = 0
    
    # Importar Empréstimos
    if "Empréstimos" in wb.sheetnames:
        ws_loans = wb["Empréstimos"]
        
        for row in ws_loans.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Pular linhas vazias
                continue
            
            # Verificar se empréstimo já existe
            existing = db.query(Emprestimo).filter(Emprestimo.id == row[0]).first()
            
            if existing:
                # Atualizar empréstimo existente
                existing.descricao = row[1]
                existing.instituicao_credora = row[2]
                existing.valor_parcela = row[3]
                existing.valor_parcela_adiantada = row[4]
                existing.qtd_total_parcelas = row[5]
                existing.qtd_parcelas_devidas = row[6]
                existing.taxa_selic_registro = row[7]
                existing.taxa_cdi_registro = row[8]
                existing.data_cadastro = row[9]
                existing.dia_vencimento = row[10]
            else:
                # Criar novo empréstimo
                new_loan = Emprestimo(
                    descricao=row[1],
                    instituicao_credora=row[2],
                    valor_parcela=row[3],
                    valor_parcela_adiantada=row[4],
                    qtd_total_parcelas=row[5],
                    qtd_parcelas_devidas=row[6],
                    taxa_selic_registro=row[7],
                    taxa_cdi_registro=row[8],
                    data_cadastro=row[9],
                    dia_vencimento=row[10]
                )
                db.add(new_loan)
            
            loans_imported += 1
    
    # Importar Histórico
    if "Histórico" in wb.sheetnames:
        ws_history = wb["Histórico"]
        
        for row in ws_history.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Pular linhas vazias
                continue
            
            # Verificar se histórico já existe
            existing = db.query(HistoricoValorAdiantado).filter(
                HistoricoValorAdiantado.id == row[0]
            ).first()
            
            if not existing:
                new_history = HistoricoValorAdiantado(
                    emprestimo_id=row[1],
                    data_registro=row[2],
                    valor_parcela_adiantada=row[3],
                    taxa_selic=row[4],
                    taxa_cdi=row[5]
                )
                db.add(new_history)
                history_imported += 1
    
    db.commit()
    
    return {
        "loans_imported": loans_imported,
        "history_imported": history_imported
    }
