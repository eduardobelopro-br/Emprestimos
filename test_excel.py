"""
Script de teste para verificar funcionalidades de exportação/importação Excel
"""
import sys
sys.path.append('backend')

from database import SessionLocal, Emprestimo, HistoricoValorAdiantado, init_db
from excel_handler import export_loans_to_excel, import_loans_from_excel, auto_sync_to_excel
from datetime import date

# Inicializar banco
init_db()

# Criar sessão
db = SessionLocal()

try:
    print("=" * 60)
    print("TESTE 1: Verificando dados existentes")
    print("=" * 60)
    
    emprestimos = db.query(Emprestimo).all()
    historicos = db.query(HistoricoValorAdiantado).all()
    
    print(f"✅ Empréstimos encontrados: {len(emprestimos)}")
    print(f"✅ Históricos encontrados: {len(historicos)}")
    
    print("\n" + "=" * 60)
    print("TESTE 2: Exportando para Excel")
    print("=" * 60)
    
    if emprestimos:
        file_path = export_loans_to_excel(emprestimos, historicos)
        print(f"✅ Arquivo Excel criado: {file_path}")
    else:
        print("⚠️  Nenhum empréstimo para exportar. Criando empréstimo de teste...")
        
        # Criar empréstimo de teste
        test_loan = Emprestimo(
            descricao="Teste Exportação",
            instituicao_credora="Banco Teste",
            valor_parcela=1000.00,
            valor_parcela_adiantada=950.00,
            qtd_total_parcelas=12,
            qtd_parcelas_devidas=10,
            taxa_selic_registro=10.50,
            taxa_cdi_registro=9.80,
            data_cadastro=date.today(),
            dia_vencimento=15
        )
        db.add(test_loan)
        db.commit()
        print("✅ Empréstimo de teste criado")
        
        # Exportar novamente
        emprestimos = db.query(Emprestimo).all()
        file_path = export_loans_to_excel(emprestimos, [])
        print(f"✅ Arquivo Excel criado: {file_path}")
    
    print("\n" + "=" * 60)
    print("TESTE 3: Sincronização automática")
    print("=" * 60)
    
    auto_sync_to_excel(db)
    print("✅ Sincronização automática executada")
    
    print("\n" + "=" * 60)
    print("TESTE 4: Verificando estrutura do arquivo Excel")
    print("=" * 60)
    
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    print(f"✅ Abas encontradas: {wb.sheetnames}")
    
    ws = wb["Empréstimos"]
    print(f"✅ Linhas na aba Empréstimos: {ws.max_row}")
    print(f"✅ Colunas na aba Empréstimos: {ws.max_column}")
    
    print("\n" + "=" * 60)
    print("✅ TODOS OS TESTES PASSARAM!")
    print("=" * 60)
    
except Exception as e:
    print(f"\n❌ ERRO: {str(e)}")
    import traceback
    traceback.print_exc()
    
finally:
    db.close()
